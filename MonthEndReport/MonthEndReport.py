# 定义月底报表相关的信息
import os
import time
from decimal import Decimal

from openpyxl.cell.read_only import EmptyCell
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import Dimension, DimensionHolder, ColumnDimension, RowDimension
from . import SolarPower
from .SolarPower import ResidualElectricityNonNaturalPerson, FullOnlineNaturalPerson, FullOnlineNonNaturalPerson, \
    ResidualElectricityNaturalPerson, SettlementInformation, HydropowerTotal, AgriculturalAndForestryWaste, \
    WasteIncineration

global global_self


# 月底报表结算
class MonthEndReport:
    # 剩余电量非自然人清单
    residualElectricityNonNaturalPersonList = []
    # 剩余电量非自然人清单
    residualElectricityNaturalPersonList = []
    # 全额上网自然人
    fullOnlineNaturalPersonList = []
    # 全额上网非自然人
    fullOnlineNonNaturalPersonList = []
    # 月份
    month = 0
    # 年份
    year = 0
    # 结算信息
    settlementInformationList = []

    # 封装 分布式结算信息map
    settlementInformationAccountMapList = {}
    settlementInformationNameMapList = {}
    settlementInformationNameAccountMapList = {}

    # 水电
    hydropowerTotalDataList = {}

    # 农林废弃物
    AgriculturalAndForestryWasteList = {}

    # 垃圾焚烧
    WasteIncinerationList = {}


def get_column():
    column = 4 + (int(MonthEndReport.month) - 1) * 3
    return column


# 定义上年累计对象
class LastYearObj:
    name = ""
    last_year_power_purchase = 0.0
    last_year_tax_included = 0.0
    last_year_tax_excluding = 0.0


def clear_object():
    # 剩余电量非自然人清单
    MonthEndReport.residualElectricityNonNaturalPersonList = []
    # 剩余电量非自然人清单
    MonthEndReport.residualElectricityNaturalPersonList = []
    # 全额上网自然人
    MonthEndReport.fullOnlineNaturalPersonList = []
    # 全额上网非自然人
    MonthEndReport.fullOnlineNonNaturalPersonList = []
    # 月份
    MonthEndReport.month = 0
    # 年份
    MonthEndReport.year = 0
    # 结算信息
    MonthEndReport.settlementInformationList = []

    # 封装 分布式结算信息map
    MonthEndReport.settlementInformationAccountMapList = {}
    MonthEndReport.settlementInformationNameMapList = {}
    MonthEndReport.settlementInformationNameAccountMapList = {}

    # 水电
    MonthEndReport.hydropowerTotalDataList = {}

    # 农林废弃物
    MonthEndReport.AgriculturalAndForestryWasteList = {}

    # 垃圾焚烧
    MonthEndReport.WasteIncinerationList = {}


# 开始处理文件
def start(self, path, wx):
    global global_self
    global_self = self
    set_m_gauge_value(self, 30)
    print("开始读取文件")

    if len(path) == 0:
        print("未选择目录程序结束")
        return
    try:
        # 获取所有文件
        list_file_name = os.listdir(path)
        for file_name in list_file_name:
            if file_name.endswith(".xls"):
                prompt_box(wx, "错误", "请检查 " + file_name + " 文件格式是否正确,希望是.xlsx")
                set_m_gauge_value(self, 0)
                return

    except OSError:
        prompt_box(wx, "提示", "路径不正确")
        return

    set_m_gauge_value(self, 35)

    #   todo 1.先处理分布式结算信息预算
    name_map_list, name_account_map_list = \
        read_power_electricity_fees(path, list_file_name)
    #
    set_m_gauge_value(self, 40)
    #
    #   todo 2.写入手工表
    write_manual_table(path, list_file_name, name_map_list, name_account_map_list)
    #
    set_m_gauge_value(self, 45)
    #
    # # todo 3.读取手工表
    read_manual_table(path, list_file_name)

    set_m_gauge_value(self, 60)
    # todo 4.读取购电费结算表-累计表
    read_hydropower_total(path, list_file_name, wx)

    set_m_gauge_value(self, 70)
    # todo 5.更改顺序
    read_hydropower_sort(path, list_file_name, wx)

    set_m_gauge_value(self, 85)

    # TODO 6.读取上年累计.
    last_year_map_list = read_last_year_tatle(path, list_file_name)
    set_m_gauge_value(self, 90)

    # todo 7.写入水电的 农场废弃物  汪清凯迪绿色能源开发有限公司 垃圾焚烧的 只有二家 延吉天楹环保能源有限公司   敦化市中能环保电力有限公司
    write_end_table(path, list_file_name, last_year_map_list)
    time.sleep(1)
    set_m_gauge_value(self, 100)
    clear_object()
    print("全部数据写入成功")
    prompt_box(wx, "已完成,要开心哦", "表格已经做好了,有什么需要跟我说哦.么么哒")
    set_m_gauge_value(self, 0)


# 读取上年累计表格,以及样式
def read_last_year_tatle(path, list_file_name):
    sheet, manual_table_name, workbook = get_workbook_sheet(list_file_name, path, "历史表样", "二级市场购电执行情况表",
                                                            True, True)
    sheet: Worksheet
    max_row = sheet.max_row
    last_year_map_list = {}

    for row in range(14, max_row):
        obj = LastYearObj()
        obj.name = sheet["B" + str(row)].value
        obj.last_year_power_purchase = sheet["K" + str(row)].value
        obj.last_year_tax_included = sheet["Q" + str(row)].value
        obj.last_year_tax_excluding = sheet["T" + str(row)].value
        last_year_map_list[sheet["B" + str(row)].value.strip()] = obj
        value: str = sheet.cell(row, 2).value
        if value.find("其他能源含从公司系统外购电") > -1:
            break

    workbook.close()
    return last_year_map_list


def set_m_gauge_value(global_self, x):
    global_self.m_gauge_进度条.SetValue(x)


# 更新水电用户数据
def read_hydropower_sort(path, file_list, wx):
    sheet, manual_table_name, workbook = get_workbook_sheet(file_list, path, "报表工作表", "排序", True, False)

    col_e_data = []

    for row in sheet.iter_rows(min_row=2, min_col=5, max_col=5, max_row=sheet.max_row, values_only=True):
        if row[0] is not None:
            col_e_data.append(row[0])
    # print(col_e_data)
    # 使用字典推导式，将MonthEndReport.hydropowerTotalDataList按照col_e_data顺序排列
    MonthEndReport.hydropowerTotalDataList = {key: MonthEndReport.hydropowerTotalDataList.get(key, None) for key in
                                              col_e_data}
    workbook.close()
    # 输出按照col_e_data顺序排列后的字典
    # print(MonthEndReport.hydropowerTotalDataList)


# 写入最终表
def write_end_table(path, list_file_name, last_year_map_list):
    route, manual_table_name = get_file_path(path, list_file_name, "空白表样")
    print(route)
    print(manual_table_name)
    workbook = load_workbook(route, read_only=False, data_only=True)
    sheet_name = get_sheet_name_by_workbook(workbook, "二级市场购电执行情况表")
    sheet: Worksheet = workbook.get_sheet_by_name(sheet_name)

    max_row = sheet.max_row + 1
    current_status = ""
    prefix = "  "
    for row in range(14, max_row):
        cell = sheet.cell(row, 2)
        if cell.value is not None:
            cell_value: str = cell.value

            if cell_value.find("3.水电") > -1:
                current_status = "水电"
                continue
            elif cell_value.find("农林废弃物") > -1:

                map_len = len(MonthEndReport.AgriculturalAndForestryWasteList)
                sheet.insert_rows(row + 1, map_len)
                num = 0
                template_row = row
                rows = str(sheet.cell(row, 3).value)
                agricultural_and_forestry_waste_list = MonthEndReport.AgriculturalAndForestryWasteList
                temp_row = row + 1
                for key in agricultural_and_forestry_waste_list.keys():
                    value = agricultural_and_forestry_waste_list.get(key)
                    num += 1

                    if value is None:
                        value = HydropowerTotal(key, 0, 0, 0, "", key, "")

                    row_one = rows + "-" + str(num)
                    sheet.cell(temp_row, 2, prefix + value.name)
                    sheet.cell(temp_row, 3, row_one)
                    sheet.cell(temp_row, 4, value.sap_name)
                    sheet.cell(temp_row, 5, "其他")
                    sheet.cell(temp_row, 6, value.unit_capacity)
                    sheet.cell(temp_row, 11, float(value.power_purchase) / 10000)
                    sheet.cell(temp_row, 17, value.tax_included)
                    sheet.cell(temp_row, 20, value.tax_excluding)
                    obj = last_year_map_list.get(key, None)
                    if obj is None:
                        sheet.cell(temp_row, 12, 0)
                        sheet.cell(temp_row, 18, 0)
                        sheet.cell(temp_row, 21, 0)
                    else:
                        sheet.cell(temp_row, 12, obj.last_year_power_purchase)
                        sheet.cell(temp_row, 18, obj.last_year_tax_included)
                        sheet.cell(temp_row, 21, obj.last_year_tax_excluding)
                    copy_row_style(sheet, temp_row, template_row)

                current_status = "None"
                continue
            elif cell_value.find("垃圾焚烧") > -1:
                current_status = "垃圾焚烧"
                continue
            elif cell_value.find("自发自用，余电上网") > 1:
                current_status = "自发自用，余电上网"
                continue
            elif cell_value.find("全额上网") > -1:
                current_status = "全额上网"
                continue

            if current_status == "水电":
                # 获得值不是None的字典长度
                map_len = len([k for k in MonthEndReport.hydropowerTotalDataList if
                               MonthEndReport.hydropowerTotalDataList[k] is not None])
                sheet.insert_rows(row, map_len)
                num = 0
                # 水电的那一行,取Font, 格式
                template_row = row - 1

                rows = str(sheet.cell(row - 1, 3).value)
                hydropower_total_data_list = MonthEndReport.hydropowerTotalDataList
                temp_row = row
                for key in hydropower_total_data_list.keys():
                    value = hydropower_total_data_list.get(key)
                    num += 1
                    value: HydropowerTotal
                    if value is None:
                        continue
                        # value = HydropowerTotal(key, 0, 0, 0, "", key, "")

                    row_one = rows + "-" + str(num)
                    sheet.cell(temp_row, 2, prefix + value.name)
                    sheet.cell(temp_row, 3, row_one)
                    sheet.cell(temp_row, 4, value.sap_name)

                    sheet.cell(temp_row, 5, "其他")
                    sheet.cell(temp_row, 6, value.unit_capacity)
                    sheet.cell(temp_row, 11, float(value.power_purchase) / 10000)
                    sheet.cell(temp_row, 17, value.tax_included)
                    sheet.cell(temp_row, 20, value.tax_excluding)
                    obj = last_year_map_list.get(key, None)
                    if obj is None:
                        sheet.cell(temp_row, 12, 0)
                        sheet.cell(temp_row, 18, 0)
                        sheet.cell(temp_row, 21, 0)
                    else:
                        sheet.cell(temp_row, 12, obj.last_year_power_purchase)
                        sheet.cell(temp_row, 18, obj.last_year_tax_included)
                        sheet.cell(temp_row, 21, obj.last_year_tax_excluding)
                    copy_row_style(sheet, temp_row, template_row)
                    temp_row += 1
                current_status = "None"

            elif current_status == "垃圾焚烧":
                map_len = len(MonthEndReport.WasteIncinerationList)
                sheet.insert_rows(row, map_len)
                num = 0
                template_row = row - 1
                rows = str(sheet.cell(row - 1, 3).value)
                waste_incineration_list = MonthEndReport.WasteIncinerationList
                temp_row = row
                for key in waste_incineration_list.keys():
                    value = waste_incineration_list.get(key)
                    num += 1
                    if value is None:
                        value = HydropowerTotal(key, 0, 0, 0, "", key, "")

                    row_one = rows + "-" + str(num)
                    sheet.cell(temp_row, 2, prefix + value.name)
                    sheet.cell(temp_row, 3, row_one)
                    sheet.cell(temp_row, 4, value.sap_name)
                    sheet.cell(temp_row, 5, "其他")
                    sheet.cell(temp_row, 6, value.unit_capacity)
                    sheet.cell(temp_row, 11, float(value.power_purchase) / 10000)
                    sheet.cell(temp_row, 17, value.tax_included)
                    sheet.cell(temp_row, 20, value.tax_excluding)
                    obj = last_year_map_list.get(key, None)
                    if obj is None:
                        sheet.cell(temp_row, 12, 0)
                        sheet.cell(temp_row, 18, 0)
                        sheet.cell(temp_row, 21, 0)
                    else:
                        sheet.cell(temp_row, 12, obj.last_year_power_purchase)
                        sheet.cell(temp_row, 18, obj.last_year_tax_included)
                        sheet.cell(temp_row, 21, obj.last_year_tax_excluding)
                    copy_row_style(sheet, temp_row, template_row)
                    temp_row += 1
                current_status = "None"

                pass
            elif current_status == "自发自用，余电上网":
                # 当前行是 其中自然人
                map_len = len(MonthEndReport.residualElectricityNaturalPersonList)
                sheet.insert_rows(row + 1, map_len)
                num = 0
                template_row = row
                rows = str(sheet.cell(row, 3).value)
                residual_electricity_natural_person_list = MonthEndReport.residualElectricityNaturalPersonList
                temp_row = row + 1
                for value in residual_electricity_natural_person_list:
                    num += 1
                    row_one = rows + "-" + str(num)
                    sheet.cell(temp_row, 2, prefix + value.name)
                    sheet.cell(temp_row, 3, row_one)
                    sheet.cell(temp_row, 4, value.sap_name)
                    sheet.cell(temp_row, 5, "其他")
                    sheet.cell(temp_row, 6, value.unit_capacity)
                    sheet.cell(temp_row, 11, value.power_purchase)
                    sheet.cell(temp_row, 17, value.tax_included)
                    sheet.cell(temp_row, 20, value.tax_excluding)
                    obj = last_year_map_list.get(value.name, None)
                    if obj is None:
                        sheet.cell(temp_row, 12, 0)
                        sheet.cell(temp_row, 18, 0)
                        sheet.cell(temp_row, 21, 0)
                    else:
                        sheet.cell(temp_row, 12, obj.last_year_power_purchase)
                        sheet.cell(temp_row, 18, obj.last_year_tax_included)
                        sheet.cell(temp_row, 21, obj.last_year_tax_excluding)
                    copy_row_style(sheet, temp_row, template_row)
                    temp_row += 1

                # 当前行是非自然人

                sheet.row_dimensions[temp_row].height = sheet.row_dimensions[template_row].height
                map_len = len(MonthEndReport.residualElectricityNonNaturalPersonList)
                sheet.insert_rows(temp_row + 1, map_len)

                num = 0
                rows = str(sheet.cell(temp_row, 3).value)
                temp_row += 1
                residual_electricity_non_natural_person_list = \
                    MonthEndReport.residualElectricityNonNaturalPersonList

                for value in residual_electricity_non_natural_person_list:
                    num += 1
                    row_one = rows + "-" + str(num)
                    sheet.cell(temp_row, 2, prefix + value.name)
                    sheet.cell(temp_row, 3, row_one)
                    sheet.cell(temp_row, 4, value.sap_name)

                    sheet.cell(temp_row, 5, "其他")
                    sheet.cell(temp_row, 6, value.unit_capacity)
                    sheet.cell(temp_row, 11, value.power_purchase)
                    sheet.cell(temp_row, 17, value.tax_included)
                    sheet.cell(temp_row, 20, value.tax_excluding)
                    obj = last_year_map_list.get(value.name, None)
                    if obj is None:
                        sheet.cell(temp_row, 12, 0)
                        sheet.cell(temp_row, 18, 0)
                        sheet.cell(temp_row, 21, 0)
                    else:
                        sheet.cell(temp_row, 12, obj.last_year_power_purchase)
                        sheet.cell(temp_row, 18, obj.last_year_tax_included)
                        sheet.cell(temp_row, 21, obj.last_year_tax_excluding)
                    copy_row_style(sheet, temp_row, template_row)
                    temp_row += 1
                sheet.row_dimensions[temp_row].height = sheet.row_dimensions[template_row].height
                current_status = "None"

            elif current_status == "全额上网":
                sheet.row_dimensions[row].height = sheet.row_dimensions[row - 2].height
                sheet.row_dimensions[row - 1].height = sheet.row_dimensions[row - 2].height

                # 当前行是 其中自然人
                map_len = len(MonthEndReport.fullOnlineNaturalPersonList)
                sheet.insert_rows(row + 1, map_len)
                num = 0
                template_row = row
                rows = str(sheet.cell(row, 3).value)
                full_online_natural_person_list = MonthEndReport.fullOnlineNaturalPersonList
                temp_row = row + 1
                for value in full_online_natural_person_list:
                    num += 1
                    row_one = rows + "-" + str(num)
                    sheet.cell(temp_row, 2, prefix + value.name)
                    sheet.cell(temp_row, 3, row_one)
                    sheet.cell(temp_row, 4, value.sap_name)
                    sheet.cell(temp_row, 5, "其他")
                    sheet.cell(temp_row, 6, value.unit_capacity)
                    sheet.cell(temp_row, 11, value.power_purchase)
                    sheet.cell(temp_row, 17, value.tax_included)
                    sheet.cell(temp_row, 20, value.tax_excluding)
                    obj = last_year_map_list.get(value.name, None)
                    if obj is None:
                        sheet.cell(temp_row, 12, 0)
                        sheet.cell(temp_row, 18, 0)
                        sheet.cell(temp_row, 21, 0)
                    else:
                        sheet.cell(temp_row, 12, obj.last_year_power_purchase)
                        sheet.cell(temp_row, 18, obj.last_year_tax_included)
                        sheet.cell(temp_row, 21, obj.last_year_tax_excluding)
                    copy_row_style(sheet, temp_row, template_row)
                    temp_row += 1
                sheet.row_dimensions[temp_row].height = sheet.row_dimensions[template_row].height
                # 当前行是非自然人
                map_len = len(MonthEndReport.fullOnlineNonNaturalPersonList)
                sheet.insert_rows(temp_row + 1, map_len)
                num = 0
                rows = str(sheet.cell(temp_row, 3).value)
                temp_row += 1
                full_online_non_natural_person_list = MonthEndReport.fullOnlineNonNaturalPersonList
                for value in full_online_non_natural_person_list:
                    num += 1
                    row_one = rows + "-" + str(num)
                    sheet.cell(temp_row, 2, prefix + value.name)
                    sheet.cell(temp_row, 3, row_one)
                    sheet.cell(temp_row, 4, value.sap_name)
                    sheet.cell(temp_row, 5, "其他")
                    sheet.cell(temp_row, 6, value.unit_capacity)
                    sheet.cell(temp_row, 11, value.power_purchase)
                    sheet.cell(temp_row, 17, value.tax_included)
                    sheet.cell(temp_row, 20, value.tax_excluding)
                    obj = last_year_map_list.get(value.name, None)
                    if obj is None:
                        sheet.cell(temp_row, 12, 0)
                        sheet.cell(temp_row, 18, 0)
                        sheet.cell(temp_row, 21, 0)
                    else:
                        sheet.cell(temp_row, 12, obj.last_year_power_purchase)
                        sheet.cell(temp_row, 18, obj.last_year_tax_included)
                        sheet.cell(temp_row, 21, obj.last_year_tax_excluding)
                    copy_row_style(sheet, temp_row, template_row)
                    temp_row += 1
                sheet.row_dimensions[temp_row].height = sheet.row_dimensions[template_row].height
                break
    workbook.save(route)
    workbook.close()


# 按行复制单元格格式
def copy_row_style(sheet, temp_row, template_row):
    sheet: Worksheet
    for col in range(1, sheet.max_column + 1):
        template_cell = sheet.cell(row=template_row, column=col)
        new_cell = sheet.cell(row=temp_row, column=col)
        new_cell.border = template_cell.border.copy()
        new_cell.fill = template_cell.fill.copy()
        new_cell.font = template_cell.font.copy()
        new_cell.number_format = template_cell.number_format
        new_cell.protection = template_cell.protection.copy()
        new_cell.alignment = template_cell.alignment.copy()
    sheet.row_dimensions[temp_row].height = sheet.row_dimensions[9].height


# 读取水电累计表
def read_hydropower_total(path, list_file_name, wx):
    sheet, manual_table_name, workbook = get_workbook_sheet(list_file_name, path, "购电费结算表", "累计", True, True)
    max_row = sheet.max_row + 1
    status = True
    for row in range(5, max_row):
        cell_one = sheet.cell(row, 1)
        cell_two = sheet.cell(row, 2)
        if status:
            if str(cell_one.value).strip() != "水电合计" and str(cell_two.value).strip() != "None":
                unit_capacity = ""
                if sheet.cell(row, 3).value is not None:
                    unit_capacity = float(sheet.cell(row, 3).value)
                # 存储水电对象
                hydropower_total = HydropowerTotal(str(sheet.cell(row, 2).value), sheet.cell(row, 4).value,
                                                   sheet.cell(row, 10).value, sheet.cell(row, 8).value,
                                                   unit_capacity, str(sheet.cell(row, 2).value), "")

                MonthEndReport.hydropowerTotalDataList[str(sheet.cell(row, 2).value)] = hydropower_total
            else:
                continue
            if str(cell_one.value).strip() == "水电合计":
                status = False
        if str(cell_one.value).strip() == "汪清凯迪绿色能源开发有限公司":

            # AgriculturalAndForestryWasteList 农林废弃物
            agricultural_and_forestry_waste = AgriculturalAndForestryWaste(str(sheet.cell(row, 2).value),
                                                                           sheet.cell(row, 4).value,
                                                                           sheet.cell(row, 10).value,
                                                                           sheet.cell(row, 8).value,
                                                                           float(sheet.cell(row, 3).value),
                                                                           str(sheet.cell(row, 2).value), "")

            MonthEndReport.AgriculturalAndForestryWasteList[
                str(sheet.cell(row, 2).value)] = agricultural_and_forestry_waste
        elif str(cell_one.value).strip() == "延吉天楹垃圾电站":

            # 垃圾焚烧 WasteIncinerationList
            power_purchase = sheet.cell(row, 4).value
            waste_incineration = WasteIncineration(str(sheet.cell(row, 2).value),
                                                   power_purchase,
                                                   sheet.cell(row, 10).value,
                                                   sheet.cell(row, 8).value,
                                                   float(sheet.cell(row, 3).value), str(sheet.cell(row, 2).value),
                                                   "")
            MonthEndReport.WasteIncinerationList[
                str(sheet.cell(row, 2).value)] = waste_incineration
        elif str(cell_one.value).strip() == "敦化中能垃圾发电厂":

            # 垃圾焚烧 WasteIncinerationList
            waste_incineration = WasteIncineration(str(sheet.cell(row, 2).value),
                                                   sheet.cell(row, 4).value,
                                                   sheet.cell(row, 10).value,
                                                   sheet.cell(row, 8).value,
                                                   float(sheet.cell(row, 3).value), str(sheet.cell(row, 2).value),
                                                   "")
            MonthEndReport.WasteIncinerationList[
                str(sheet.cell(row, 2).value)] = waste_incineration
            break
    workbook.close()


# 封装 根据文件名,sheet名,获取表格操作对象
def get_workbook_sheet(list_file_name, path, file_name, sheet_name, read_only, data_only):
    route, manual_table_name = get_file_path(path, list_file_name, file_name)
    workbook = load_workbook(route, read_only=read_only, data_only=data_only, )
    sheet_name = get_sheet_name_by_workbook(workbook, sheet_name)
    sheet = workbook.get_sheet_by_name(sheet_name)
    return sheet, manual_table_name, workbook


# 写入手工表
def write_manual_table(path, file_list, name_map_list, name_account_map_list):
    route, manual_table_name = get_file_path(path, file_list, "手工表")
    workbook = load_workbook(route, read_only=False, data_only=True)
    sheet_name = get_sheet_name_by_workbook(workbook, MonthEndReport.year)
    sheet = workbook.get_sheet_by_name(sheet_name)

    max_row = sheet.max_row + 1

    for row in range(1, max_row):
        cell = sheet.cell(row, 2)
        value = str(cell.value).strip()
        # print("准备遍历的名字是: {}", value)

        if value in name_map_list:
            status = "name_map_list"
        elif value in name_account_map_list:
            status = "name_account_map_list"
        else:
            continue
        if status == "name_map_list":
            write_data(name_map_list, row, sheet, value)

        elif status == "name_account_map_list":
            write_data(name_account_map_list, row, sheet, value)

    workbook.save(route)
    workbook.close()


def write_data(data_list, row, sheet, value):
    user: SolarPower.SettlementInformation = data_list.get(value)
    sheet.cell(row, get_column(), float(user.power_purchase) / 10000)
    sheet.cell(row, get_column() + 1, user.tax_included)
    sheet.cell(row, get_column() + 2, user.tax_excluding)
    total_power_purchase = 0.00
    total_tax_included = 0.00
    total_tax_excluding = 0.00
    for i in range(1, 13):
        column = 4 + (i - 1) * 3
        cell = sheet.cell(row, column)
        if cell.value is not None:
            total_power_purchase += float(cell.value)
        else:
            total_power_purchase += 0
        column = 5 + (i - 1) * 3
        cell = sheet.cell(row, column)
        if cell.value is not None:
            total_tax_included += float(cell.value)
        else:
            total_tax_included += 0
        column = 6 + (i - 1) * 3
        cell = sheet.cell(row, column)
        if cell.value is not None:
            total_tax_excluding += float(cell.value)
        else:
            total_tax_excluding += 0

    sheet["AQ" + str(row)].value = total_power_purchase
    sheet["AR" + str(row)].value = total_tax_included
    sheet["AS" + str(row)].value = total_tax_excluding


# 读取手工表 并写入公式 获取太阳能分布式发点用户
def read_manual_table(path, file_list):
    route, manual_table_name = get_file_path(path, file_list, "手工表")

    workbook = load_workbook(route, data_only=False)
    sheet_name = get_sheet_name_by_workbook(workbook, MonthEndReport.year)
    sheet: Worksheet = workbook.get_sheet_by_name(sheet_name)
    max_row = sheet.max_row + 1

    task_status = ""
    for row in range(1, max_row):

        cell = sheet.cell(row, 2)

        if cell.value == "余电上网-非自然人":
            # print("余电上网-非自然人")
            task_status = "余电上网-非自然人"
            continue
        if cell.value is None:
            # print("空格换行")
            continue
        if cell.value == "全额上网-非自然人":
            # print("全额上网-非自然人")
            task_status = "全额上网-非自然人"
            continue
        if cell.value == "余电上网-自然人":
            # print("余电上网-自然人")
            task_status = "余电上网-自然人"
            continue
        if cell.value == "全额上网-自然人":
            # print("全额上网-自然人")
            task_status = "全额上网-自然人"
            continue

        if task_status == "余电上网-自然人":
            account, sap_name = get_name_account(cell)
            MonthEndReport.residualElectricityNaturalPersonList.append(
                ResidualElectricityNaturalPerson(cell.value.strip(), sheet["AQ" + str(row)].value,
                                                 sheet["AR" + str(row)].value,
                                                 sheet["AS" + str(row)].value, sheet["C" + str(row)].value,
                                                 sap_name, account))
            # 写入公式
            write_formula(row, sheet)

        elif task_status == "余电上网-非自然人":
            account, sap_name = get_name_account(cell)
            MonthEndReport.residualElectricityNonNaturalPersonList.append(
                ResidualElectricityNonNaturalPerson(cell.value.strip(), sheet["AQ" + str(row)].value,
                                                    sheet["AR" + str(row)].value,
                                                    sheet["AS" + str(row)].value, sheet["C" + str(row)].value,
                                                    sap_name, account))
            # 写入公式
            write_formula(row, sheet)
        elif task_status == "全额上网-自然人":
            account, sap_name = get_name_account(cell)
            MonthEndReport.fullOnlineNaturalPersonList.append(
                FullOnlineNaturalPerson(cell.value.strip(), sheet["AQ" + str(row)].value,
                                        sheet["AR" + str(row)].value,
                                        sheet["AS" + str(row)].value, sheet["C" + str(row)].value,
                                        sap_name, account))
            # 写入公式
            write_formula(row, sheet)
        elif task_status == "全额上网-非自然人":
            account, sap_name = get_name_account(cell)
            MonthEndReport.fullOnlineNonNaturalPersonList.append(
                FullOnlineNonNaturalPerson(cell.value.strip(), sheet["AQ" + str(row)].value,
                                           sheet["AR" + str(row)].value,
                                           sheet["AS" + str(row)].value, sheet["C" + str(row)].value,
                                           sap_name, account))
            # 写入公式
            write_formula(row, sheet)
    workbook.save(route)
    workbook.close()


# 写入公式
def write_formula(row, sheet):
    formula = ["D", "G", "J", "M", "P", "S", "V", "Y", "AB", "AE", "AH", "AK"]
    join = "=" + "+".join([x + str(row) for x in formula])
    sheet["AN" + str(row)].value = join

    formula = ["E", "H", "K", "N", "Q", "T", "W", "Z", "AC", "AF", "AI", "AL"]
    join = "=" + "+".join([x + str(row) for x in formula])
    sheet["AO" + str(row)].value = join

    formula = ["F", "I", "L", "O", "R", "U", "X", "AA", "AD", "AG", "AJ", "AM"]
    join = "=" + "+".join([x + str(row) for x in formula])
    sheet["AP" + str(row)].value = join


# 读取电量电费表,获得当前月份最终数据
def read_power_electricity_fees(path, file_list):
    sheet, file_name, workbook = get_workbook_sheet(file_list, path, "分布式结算信息", "电量电费", True, True)
    MonthEndReport.month = file_name[-9:-7]
    MonthEndReport.year = file_name[-13:-9]
    print("MonthEndReport.month=", MonthEndReport.month)
    print("MonthEndReport.year=", MonthEndReport.year)

    max_row = sheet.max_row

    yellow_fill: PatternFill = PatternFill("solid", fgColor="FFFFFF00", bgColor="FFFFFF00")
    orange_fill: PatternFill = PatternFill("solid", fgColor="FFFFC000", bgColor="FFFFC000")
    # green_fill: PatternFill = PatternFill("solid", fgColor="FF92D050", bgColor="FF92D050")
    # 定义公共对象
    information = SettlementInformation("", 0, 0, 0, "", "", "")
    is_object_start = True
    for row in range(1, max_row + 1):

        cell: EmptyCell = sheet.cell(row, 1)
        if cell.value is None:
            continue
        fill: PatternFill = cell.fill
        if cell.fill is not None:
            # 判断是不是标记黄色
            if fill.fgColor.rgb == yellow_fill.fgColor.rgb:
                # 如果对象开始 保存 name  account
                if is_object_start:
                    information.name = cell.value
                    information.sap_name = cell.value
                    cell: EmptyCell = sheet.cell(row, 2)
                    information.account = cell.value
                    is_object_start = False
                else:
                    cell: EmptyCell = sheet.cell(row, 1)
                    if cell.value == "电费年月":
                        continue
                    # 不是开始,
                    cell_d: EmptyCell = sheet.cell(row, 4)
                    d_fill: PatternFill = cell_d.fill
                    if d_fill is not None:
                        if d_fill.fgColor.rgb == orange_fill.fgColor.rgb:
                            information.power_purchase = Decimal(sheet["B" + str(row)].internal_value) + Decimal(
                                information.power_purchase)
                            information.tax_included = Decimal(sheet["D" + str(row)].internal_value) + Decimal(
                                information.tax_included)
                            information.tax_excluding = Decimal(sheet["F" + str(row)].internal_value) + Decimal(
                                information.tax_excluding)
                            # 如果下一行是黄色 全是数字,证明未结束,继续处理,否则本次对象结束
                            cell: EmptyCell = sheet.cell(row + 1, 1)

                            if cell.value is not None and str(cell.value).isdigit() \
                                    and cell.fill.fgColor.rgb == yellow_fill.fgColor.rgb:
                                is_object_start = False
                                continue
                            else:
                                settlement_information = SettlementInformation(information.name,
                                                                               format(information.power_purchase,
                                                                                      '.2f'),
                                                                               format(information.tax_included, '.2f'),
                                                                               format(information.tax_excluding, '.2f'),
                                                                               "",
                                                                               information.sap_name,
                                                                               information.account)
                                MonthEndReport.settlementInformationList.append(settlement_information)
                                information.name = ""
                                information.power_purchase = 0
                                information.tax_included = 0
                                information.tax_excluding = 0
                                information.sap_name = ""
                                information.account = ""
                                is_object_start = True
                        else:
                            continue
    workbook.close()

    for item in MonthEndReport.settlementInformationList:
        MonthEndReport.settlementInformationNameMapList[item.sap_name] = item
        name_account = item.sap_name + str(item.account)
        MonthEndReport.settlementInformationNameAccountMapList[name_account] = item

    # print(len(MonthEndReport.settlementInformationList))
    # print(str(MonthEndReport.settlementInformationAccountMapList))
    # print(str(MonthEndReport.settlementInformationNameMapList))
    # print(str(MonthEndReport.settlementInformationNameAccountMapList))
    # MonthEndReport.clear_object(None)
    return MonthEndReport.settlementInformationNameMapList, \
        MonthEndReport.settlementInformationNameAccountMapList


# 根据包含的名字获取sheet名字
def get_sheet_name_by_workbook(workbook, name):
    # 查看所有工作表
    sheet_names = workbook.sheetnames
    # print("查看所有工作表", sheet_names)
    work_sheet_name = ""
    # 遍历sheet
    for i in sheet_names:
        if i.__contains__(name):
            work_sheet_name = i
    # print("输出工作sheet名字\t" + work_sheet_name)
    return work_sheet_name


def prompt_box(wx, title, news):
    # 创建提示对话框
    dlg = wx.MessageDialog(None, news, title, wx.OK)
    # 显示对话框
    dlg.ShowModal()
    # 关闭对话框
    dlg.Destroy()


# 获取名字 与户号
def get_name_account(cell):
    sap_name = ""
    account = ""
    for char in cell.value:
        if char.isalpha():  # 判断是否为汉字
            sap_name += char
        elif char.isdigit():  # 判断是否为数字
            account += char
    # print(sap_name)  # 输出汉字
    # print(account)  # 输出数字
    return account, sap_name


# 获取文件路径
def get_file_path(path, file_list, name):
    manual_table_name = ""
    for file_name in file_list:
        if str(file_name).count(name) > 0:
            manual_table_name = file_name

    route = path + "\\" + manual_table_name
    return route, manual_table_name
